#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ガントチャート生成スクリプト（JTC業務品質版）
==============================================
特徴:
  - 日本の祝日に対応（振替休日含む、2025-2027年）
  - 進捗率・ステータス・優先度・担当者・備考を管理
  - 稼働日数を自動計算（土日祝を除外）
  - 曜日表示（月火水木金土日）
  - 週末＝薄グレー、祝日＝薄ピンクで色分け
  - マイルストーン対応（開始日＝終了日のタスク）
  - 月の区切り線を太線で表示
  - タスク依存関係（depends_on）対応
  - JSON/CSVからのタスクインポート対応
  - CLIオプション（出力ファイル名、インポート元指定）
  - 凡例・使い方シート付き
  - 印刷設定（A3横向き、縮小印刷対応）

使い方:
  1. 下の TASKS リストを編集（または外部JSON/CSVを用意）
  2. pip install openpyxl
  3. python ガントチャート生成.py
  4. オプション: python ガントチャート生成.py --import tasks.json -o output.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import argparse
import csv
import json
import math
import os
import sys

# ==================================================
#  ★ タスクデータ（ここを編集してください）★
# ==================================================
TASKS = [
    # Phase 1: 企画
    {"id": 1, "name": "要件定義",           "start": date(2026, 4,  1), "end": date(2026, 4,  7),
     "assignee": "田中",  "status": "完了",   "priority": "高", "progress": 100, "note": "顧客ヒアリング完了",
     "group": "企画"},
    {"id": 2, "name": "基本設計",           "start": date(2026, 4,  6), "end": date(2026, 4, 14),
     "assignee": "鈴木",  "status": "進行中", "priority": "高", "progress": 60,  "note": "DB設計レビュー待ち",
     "group": "企画", "depends_on": [1]},
    {"id": 3, "name": "詳細設計",           "start": date(2026, 4, 13), "end": date(2026, 4, 22),
     "assignee": "佐藤",  "status": "未着手", "priority": "中", "progress": 0,   "note": "",
     "group": "企画", "depends_on": [2]},
    # Phase 2: 開発
    {"id": 4, "name": "フロントエンド開発",  "start": date(2026, 4, 20), "end": date(2026, 5,  1),
     "assignee": "高橋",  "status": "未着手", "priority": "高", "progress": 0,   "note": "React + TypeScript",
     "group": "開発", "depends_on": [3]},
    {"id": 5, "name": "バックエンド開発",    "start": date(2026, 4, 22), "end": date(2026, 5,  8),
     "assignee": "伊藤",  "status": "未着手", "priority": "高", "progress": 0,   "note": "API設計書参照",
     "group": "開発", "depends_on": [3]},
    {"id": 6, "name": "DB構築",             "start": date(2026, 4, 20), "end": date(2026, 4, 28),
     "assignee": "渡辺",  "status": "未着手", "priority": "中", "progress": 0,   "note": "PostgreSQL",
     "group": "開発", "depends_on": [3]},
    # Phase 3: テスト・リリース
    {"id": 7, "name": "結合テスト",         "start": date(2026, 5,  7), "end": date(2026, 5, 15),
     "assignee": "山本",  "status": "未着手", "priority": "高", "progress": 0,   "note": "テスト仕様書作成中",
     "group": "テスト", "depends_on": [4, 5, 6]},
    {"id": 8, "name": "受入テスト（UAT）",   "start": date(2026, 5, 14), "end": date(2026, 5, 22),
     "assignee": "中村",  "status": "未着手", "priority": "高", "progress": 0,   "note": "顧客参加",
     "group": "テスト", "depends_on": [7]},
    {"id": 9, "name": "リリース準備",       "start": date(2026, 5, 21), "end": date(2026, 5, 27),
     "assignee": "田中",  "status": "未着手", "priority": "中", "progress": 0,   "note": "インフラ・ドキュメント",
     "group": "リリース", "depends_on": [8]},
    {"id": 10, "name": "★ 本番リリース",     "start": date(2026, 5, 28), "end": date(2026, 5, 28),
     "assignee": "全員",  "status": "未着手", "priority": "最高", "progress": 0,  "note": "マイルストーン",
     "group": "リリース", "depends_on": [9]},
]

PROJECT_TITLE = "○○システム開発プロジェクト 工程表"
OUTPUT_FILE = "ガントチャート.xlsx"

# ==================================================
#  日本の祝日（2025-2027年、振替休日含む）
# ==================================================
HOLIDAYS = {
    # 2025年
    date(2025,  1,  1): "元日",
    date(2025,  1, 13): "成人の日",
    date(2025,  2, 11): "建国記念の日",
    date(2025,  2, 23): "天皇誕生日",
    date(2025,  2, 24): "振替休日",
    date(2025,  3, 20): "春分の日",
    date(2025,  4, 29): "昭和の日",
    date(2025,  5,  3): "憲法記念日",
    date(2025,  5,  4): "みどりの日",
    date(2025,  5,  5): "こどもの日",
    date(2025,  5,  6): "振替休日",
    date(2025,  7, 21): "海の日",
    date(2025,  8, 11): "山の日",
    date(2025,  9, 15): "敬老の日",
    date(2025,  9, 23): "秋分の日",
    date(2025, 10, 13): "スポーツの日",
    date(2025, 11,  3): "文化の日",
    date(2025, 11, 23): "勤労感謝の日",
    date(2025, 11, 24): "振替休日",
    # 2026年
    date(2026,  1,  1): "元日",
    date(2026,  1, 12): "成人の日",
    date(2026,  2, 11): "建国記念の日",
    date(2026,  2, 23): "天皇誕生日",
    date(2026,  3, 20): "春分の日",
    date(2026,  4, 29): "昭和の日",
    date(2026,  5,  3): "憲法記念日",
    date(2026,  5,  4): "みどりの日",
    date(2026,  5,  5): "こどもの日",
    date(2026,  5,  6): "振替休日",
    date(2026,  7, 20): "海の日",
    date(2026,  8, 11): "山の日",
    date(2026,  9, 21): "敬老の日",
    date(2026,  9, 23): "秋分の日",
    date(2026, 10, 12): "スポーツの日",
    date(2026, 11,  3): "文化の日",
    date(2026, 11, 23): "勤労感謝の日",
    # 2027年
    date(2027,  1,  1): "元日",
    date(2027,  1, 11): "成人の日",
    date(2027,  2, 11): "建国記念の日",
    date(2027,  2, 23): "天皇誕生日",
    date(2027,  3, 21): "春分の日",
    date(2027,  3, 22): "振替休日",
    date(2027,  4, 29): "昭和の日",
    date(2027,  5,  3): "憲法記念日",
    date(2027,  5,  4): "みどりの日",
    date(2027,  5,  5): "こどもの日",
    date(2027,  7, 19): "海の日",
    date(2027,  8, 11): "山の日",
    date(2027,  9, 20): "敬老の日",
    date(2027,  9, 23): "秋分の日",
    date(2027, 10, 11): "スポーツの日",
    date(2027, 11,  3): "文化の日",
    date(2027, 11, 23): "勤労感謝の日",
}

# 後方互換性のためのエイリアス
HOLIDAYS_2026 = {k: v for k, v in HOLIDAYS.items() if k.year == 2026}


def is_holiday(d):
    return d in HOLIDAYS


def is_non_working(d):
    return d.weekday() >= 5 or is_holiday(d)


def count_working_days(start, end):
    count = 0
    d = start
    while d <= end:
        if not is_non_working(d):
            count += 1
        d += timedelta(days=1)
    return count


# ==================================================
#  タスクバリデーション
# ==================================================
VALID_STATUSES = {"完了", "進行中", "未着手", "遅延", "中断"}
VALID_PRIORITIES = {"最高", "高", "中", "低"}


class TaskValidationError(Exception):
    """タスクデータのバリデーションエラー"""
    pass


def validate_task(task, index):
    """単一タスクのバリデーション。エラーがあればTaskValidationErrorを送出。"""
    errors = []
    label = f"タスク#{index + 1}"

    if "name" not in task or not task["name"]:
        errors.append(f"{label}: 'name' は必須です")

    if "start" not in task:
        errors.append(f"{label}: 'start' は必須です")
    elif not isinstance(task["start"], date):
        errors.append(f"{label}: 'start' はdate型である必要があります")

    if "end" not in task:
        errors.append(f"{label}: 'end' は必須です")
    elif not isinstance(task["end"], date):
        errors.append(f"{label}: 'end' はdate型である必要があります")

    if isinstance(task.get("start"), date) and isinstance(task.get("end"), date):
        if task["start"] > task["end"]:
            errors.append(f"{label} '{task.get('name', '?')}': 開始日({task['start']})が終了日({task['end']})より後です")

    status = task.get("status", "未着手")
    if status not in VALID_STATUSES:
        errors.append(f"{label} '{task.get('name', '?')}': 無効なステータス '{status}'。有効値: {VALID_STATUSES}")

    priority = task.get("priority", "中")
    if priority not in VALID_PRIORITIES:
        errors.append(f"{label} '{task.get('name', '?')}': 無効な優先度 '{priority}'。有効値: {VALID_PRIORITIES}")

    progress = task.get("progress", 0)
    if not isinstance(progress, (int, float)) or progress < 0 or progress > 100:
        errors.append(f"{label} '{task.get('name', '?')}': progress は0〜100の数値である必要があります（現在: {progress}）")

    if errors:
        raise TaskValidationError("\n".join(errors))


def validate_tasks(tasks):
    """全タスクをバリデーション。依存関係の整合性もチェック。"""
    all_errors = []
    task_ids = set()

    for i, task in enumerate(tasks):
        try:
            validate_task(task, i)
        except TaskValidationError as e:
            all_errors.append(str(e))

        task_id = task.get("id")
        if task_id is not None:
            if task_id in task_ids:
                all_errors.append(f"タスク#{i + 1} '{task.get('name', '?')}': ID {task_id} が重複しています")
            task_ids.add(task_id)

    # 依存関係チェック
    for i, task in enumerate(tasks):
        depends = task.get("depends_on", [])
        for dep_id in depends:
            if dep_id not in task_ids:
                all_errors.append(
                    f"タスク#{i + 1} '{task.get('name', '?')}': "
                    f"依存先ID {dep_id} が存在しません"
                )

    if all_errors:
        raise TaskValidationError("\n".join(all_errors))


# ==================================================
#  JSON/CSVインポート
# ==================================================
def import_tasks_from_json(filepath):
    """JSONファイルからタスクをインポート。"""
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)

    tasks_data = data if isinstance(data, list) else data.get("tasks", [])
    tasks = []
    for item in tasks_data:
        task = dict(item)
        task["start"] = date.fromisoformat(task["start"])
        task["end"] = date.fromisoformat(task["end"])
        task["progress"] = int(task.get("progress", 0))
        if "depends_on" in task and isinstance(task["depends_on"], str):
            task["depends_on"] = [int(x.strip()) for x in task["depends_on"].split(",") if x.strip()]
        tasks.append(task)
    return tasks


def import_tasks_from_csv(filepath):
    """CSVファイルからタスクをインポート。"""
    tasks = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            task = {
                "name": row["name"],
                "start": date.fromisoformat(row["start"]),
                "end": date.fromisoformat(row["end"]),
                "assignee": row.get("assignee", ""),
                "status": row.get("status", "未着手"),
                "priority": row.get("priority", "中"),
                "progress": int(row.get("progress", 0)),
                "note": row.get("note", ""),
                "group": row.get("group", ""),
            }
            if row.get("id"):
                task["id"] = int(row["id"])
            if row.get("depends_on"):
                task["depends_on"] = [int(x.strip()) for x in row["depends_on"].split(",") if x.strip()]
            tasks.append(task)
    return tasks


def export_tasks_to_json(tasks, filepath):
    """タスクをJSONファイルにエクスポート。"""
    export = []
    for task in tasks:
        t = dict(task)
        t["start"] = t["start"].isoformat()
        t["end"] = t["end"].isoformat()
        export.append(t)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump({"tasks": export}, f, ensure_ascii=False, indent=2)


# ==================================================
#  カラーパレット
# ==================================================
TASK_COLORS = [
    {"bar": "FF4472C4", "light": "FFD6E4F7"},  # 青
    {"bar": "FFED7D31", "light": "FFFCE4D0"},  # オレンジ
    {"bar": "FF70AD47", "light": "FFD8EDCB"},  # 緑
    {"bar": "FFFFC000", "light": "FFFFF2CC"},  # 黄
    {"bar": "FF5B9BD5", "light": "FFD3E8F5"},  # 水色
    {"bar": "FFFF6699", "light": "FFFFD6E5"},  # ピンク
    {"bar": "FF7030A0", "light": "FFE8D5F5"},  # 紫
    {"bar": "FF00B0F0", "light": "FFD0EEFB"},  # スカイ
    {"bar": "FF92D050", "light": "FFE2F0D0"},  # ライム
    {"bar": "FFFF4444", "light": "FFFFD5D5"},  # レッド
]

GROUP_COLORS = {
    "企画":     {"bg": "FF2F5496", "fg": "FFFFFFFF"},
    "開発":     {"bg": "FF4472C4", "fg": "FFFFFFFF"},
    "テスト":   {"bg": "FF70AD47", "fg": "FFFFFFFF"},
    "リリース": {"bg": "FFED7D31", "fg": "FFFFFFFF"},
}

STATUS_COLORS = {
    "完了":   {"bg": "FFE2EFDA", "fg": "FF375623"},
    "進行中": {"bg": "FFDCE6F1", "fg": "FF1F4E79"},
    "未着手": {"bg": "FFFFF2CC", "fg": "FF7F6000"},
    "遅延":   {"bg": "FFFCE4EC", "fg": "FFC00000"},
    "中断":   {"bg": "FFE8E8E8", "fg": "FF595959"},
}

PRIORITY_COLORS = {
    "最高": {"bg": "FFFF0000", "fg": "FFFFFFFF"},
    "高":   {"bg": "FFFF6600", "fg": "FFFFFFFF"},
    "中":   {"bg": "FFFFC000", "fg": "FF333333"},
    "低":   {"bg": "FF92D050", "fg": "FF333333"},
}

# ==================================================
#  スタイル定数
# ==================================================
FONT_NAME  = "Meiryo UI"
FONT_ALT   = "Arial"
HEADER_BG  = "FF1F3864"
HEADER_FG  = "FFFFFFFF"
MONTH_BG   = "FF2F5496"
SUBHEAD_BG = "FFD6DCE4"
WEEKEND_BG = "FFE8E8E8"
HOLIDAY_BG = "FFFCE4EC"
ROW_EVEN   = "FFFFFFFF"
ROW_ODD    = "FFF5F7FA"
TODAY_RED   = "FFFF0000"
MILESTONE_COLOR = "FFFF4444"
GRIDLINE   = "FFBFBFBF"
MONTH_LINE = "FF595959"
PROGRESS_GREEN = "FF70AD47"
PROGRESS_BG    = "FFD9D9D9"


def make_fill(c):
    if len(c) == 6:
        c = "FF" + c
    return PatternFill("solid", fgColor=c)


def border(top="thin", bottom="thin", left="thin", right="thin",
           top_c=GRIDLINE, bottom_c=GRIDLINE, left_c=GRIDLINE, right_c=GRIDLINE):
    def _side(style, color):
        return Side(style=style, color=color) if style else Side(style=None)
    return Border(top=_side(top, top_c), bottom=_side(bottom, bottom_c),
                  left=_side(left, left_c), right=_side(right, right_c))


def std_border():
    return border()


def font(size=10, bold=False, color="FF333333", italic=False):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color, italic=italic)


def align(h="center", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


# ==================================================
#  入力シート
# ==================================================
def create_input_sheet(wb, tasks):
    ws = wb.active
    ws.title = "入力"

    # --- プロジェクトタイトル ---
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = PROJECT_TITLE
    c.font = font(18, bold=True, color=HEADER_FG)
    c.fill = make_fill(HEADER_BG)
    c.alignment = align()
    ws.row_dimensions[1].height = 42

    # --- サブヘッダー ---
    ws.merge_cells("A2:L2")
    c = ws["A2"]
    c.value = f"作成日: {date.today().strftime('%Y年%m月%d日')}　｜　タスクを編集後「ガントチャート生成.py」を実行してガントチャートを再生成"
    c.font = font(9, italic=True, color="FF595959")
    c.fill = make_fill("FFEEF2FA")
    c.alignment = align(h="left", wrap=True)
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 6

    # --- ヘッダー行 ---
    headers = [
        ("No.",       5),
        ("グループ",   10),
        ("ステータス",  10),
        ("優先度",     8),
        ("タスク名",   28),
        ("担当者",     10),
        ("開始日",     13),
        ("終了日",     13),
        ("暦日数",     8),
        ("稼働日数",    9),
        ("進捗率",     8),
        ("備考",       30),
    ]
    HR = 4
    for col, (h, w) in enumerate(headers, start=1):
        c = ws.cell(row=HR, column=col, value=h)
        c.font = font(9, bold=True, color=HEADER_FG)
        c.fill = make_fill(MONTH_BG)
        c.alignment = align()
        c.border = std_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[HR].height = 24

    # --- データ行 ---
    for i, task in enumerate(tasks):
        row = 5 + i
        bg = ROW_EVEN if i % 2 == 0 else ROW_ODD
        fill = make_fill(bg)

        def cell(col, value, bold=False, h="center", num_fmt=None, custom_fill=None, fg_color="FF333333"):
            c = ws.cell(row=row, column=col, value=value)
            c.font = font(9, bold=bold, color=fg_color)
            c.fill = custom_fill if custom_fill else fill
            c.alignment = align(h=h, indent=1 if h == "left" else 0)
            c.border = std_border()
            if num_fmt:
                c.number_format = num_fmt
            return c

        # No.
        cell(1, i + 1)

        # グループ
        group = task.get("group", "")
        gc = GROUP_COLORS.get(group)
        if gc:
            cell(2, group, bold=True, custom_fill=make_fill(gc["bg"]), fg_color=gc["fg"])
        else:
            cell(2, group)

        # ステータス
        st = task.get("status", "未着手")
        sc = STATUS_COLORS.get(st, STATUS_COLORS["未着手"])
        cell(3, st, bold=True, custom_fill=make_fill(sc["bg"]), fg_color=sc["fg"])

        # 優先度
        pr = task.get("priority", "中")
        pc = PRIORITY_COLORS.get(pr, PRIORITY_COLORS["中"])
        cell(4, pr, bold=True, custom_fill=make_fill(pc["bg"]), fg_color=pc["fg"])

        # タスク名
        cell(5, task["name"], bold=True, h="left")

        # 担当者
        cell(6, task.get("assignee", ""))

        # 開始日
        cell(7, task["start"], num_fmt="YYYY/MM/DD")
        # 終了日
        cell(8, task["end"], num_fmt="YYYY/MM/DD")

        # 暦日数（数式）
        cell(9, f"=H{row}-G{row}+1", num_fmt='0"日"')

        # 稼働日数
        wd = count_working_days(task["start"], task["end"])
        cell(10, wd, num_fmt='0"日"', fg_color="FF2F5496")

        # 進捗率
        prog = task.get("progress", 0) / 100.0
        c = cell(11, prog, num_fmt='0%')
        if prog >= 1.0:
            c.font = font(9, bold=True, color="FF375623")
        elif prog > 0:
            c.font = font(9, bold=True, color="FF1F4E79")

        # 備考
        dep_text = ""
        depends = task.get("depends_on", [])
        if depends:
            dep_text = f"[依存: #{',#'.join(str(d) for d in depends)}] "
        cell(12, dep_text + task.get("note", ""), h="left")

        ws.row_dimensions[row].height = 22

    # --- 合計行 ---
    tr = 5 + len(tasks)
    ws.merge_cells(f"A{tr}:H{tr}")
    c = ws.cell(row=tr, column=1, value="合計タスク数 / 完了率")
    c.font = font(9, bold=True, color=HEADER_FG)
    c.fill = make_fill(HEADER_BG)
    c.alignment = align(h="right", indent=1)
    c.border = std_border()
    for col in range(2, 9):
        ws.cell(row=tr, column=col).fill = make_fill(HEADER_BG)
        ws.cell(row=tr, column=col).border = std_border()

    c = ws.cell(row=tr, column=9, value=len(tasks))
    c.font = font(9, bold=True, color=HEADER_FG)
    c.fill = make_fill(HEADER_BG)
    c.alignment = align()
    c.border = std_border()
    c.number_format = '0"件"'

    # 稼働日数合計
    c = ws.cell(row=tr, column=10, value=f"=SUM(J5:J{tr-1})")
    c.font = font(9, bold=True, color=HEADER_FG)
    c.fill = make_fill(HEADER_BG)
    c.alignment = align()
    c.border = std_border()
    c.number_format = '0"日"'

    # 平均進捗率
    c = ws.cell(row=tr, column=11, value=f"=AVERAGE(K5:K{tr-1})")
    c.font = font(9, bold=True, color=HEADER_FG)
    c.fill = make_fill(HEADER_BG)
    c.alignment = align()
    c.border = std_border()
    c.number_format = '0%'

    c = ws.cell(row=tr, column=12)
    c.fill = make_fill(HEADER_BG)
    c.border = std_border()
    ws.row_dimensions[tr].height = 24

    # --- サマリー行（グループ別進捗）---
    groups = []
    seen = set()
    for task in tasks:
        g = task.get("group", "")
        if g and g not in seen:
            groups.append(g)
            seen.add(g)

    if groups:
        sr = tr + 2
        ws.cell(row=sr, column=1, value="■ グループ別サマリー").font = font(10, bold=True, color=HEADER_BG[2:])
        ws.row_dimensions[sr].height = 20

        for col, val in enumerate(["グループ", "タスク数", "完了", "進行中", "平均進捗率"], start=1):
            c = ws.cell(row=sr + 1, column=col, value=val)
            c.font = font(9, bold=True, color=HEADER_FG)
            c.fill = make_fill(MONTH_BG)
            c.alignment = align()
            c.border = std_border()

        for idx, group in enumerate(groups):
            r = sr + 2 + idx
            group_tasks = [t for t in tasks if t.get("group") == group]
            done = sum(1 for t in group_tasks if t.get("status") == "完了")
            in_prog = sum(1 for t in group_tasks if t.get("status") == "進行中")
            avg_prog = sum(t.get("progress", 0) for t in group_tasks) / len(group_tasks) if group_tasks else 0

            bg = ROW_EVEN if idx % 2 == 0 else ROW_ODD
            fill = make_fill(bg)

            gc = GROUP_COLORS.get(group)
            if gc:
                c = ws.cell(row=r, column=1, value=group)
                c.font = font(9, bold=True, color=gc["fg"])
                c.fill = make_fill(gc["bg"])
            else:
                c = ws.cell(row=r, column=1, value=group)
                c.font = font(9, bold=True)
                c.fill = fill
            c.alignment = align()
            c.border = std_border()

            for col, val in [(2, len(group_tasks)), (3, done), (4, in_prog)]:
                c = ws.cell(row=r, column=col, value=val)
                c.font = font(9)
                c.fill = fill
                c.alignment = align()
                c.border = std_border()

            c = ws.cell(row=r, column=5, value=avg_prog / 100.0)
            c.font = font(9, bold=True)
            c.fill = fill
            c.alignment = align()
            c.border = std_border()
            c.number_format = '0%'

            ws.row_dimensions[r].height = 20

        sr = sr + 2 + len(groups) + 1
    else:
        sr = tr + 2

    # --- 祝日カレンダー ---
    hr = sr + 1
    # 対象年の祝日のみ表示
    task_years = set()
    for task in tasks:
        task_years.add(task["start"].year)
        task_years.add(task["end"].year)
    relevant_holidays = {k: v for k, v in HOLIDAYS.items() if k.year in task_years}

    ws.cell(row=hr, column=1, value=f"■ {'/'.join(str(y) for y in sorted(task_years))}年 祝日一覧").font = font(10, bold=True, color=HEADER_BG[2:])
    ws.row_dimensions[hr].height = 20

    for col, val in enumerate(["日付", "曜日", "祝日名"], start=1):
        c = ws.cell(row=hr+1, column=col, value=val)
        c.font = font(9, bold=True, color=HEADER_FG)
        c.fill = make_fill(MONTH_BG)
        c.alignment = align()
        c.border = std_border()

    sorted_holidays = sorted(relevant_holidays.items())
    WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]
    for idx, (hd, hname) in enumerate(sorted_holidays):
        r = hr + 2 + idx
        bg = ROW_EVEN if idx % 2 == 0 else ROW_ODD
        fill = make_fill(bg)

        c = ws.cell(row=r, column=1, value=hd)
        c.font = font(9); c.fill = fill; c.alignment = align()
        c.border = std_border(); c.number_format = "YYYY/MM/DD"

        c = ws.cell(row=r, column=2, value=WEEKDAY_JP[hd.weekday()])
        c.font = font(9); c.fill = fill; c.alignment = align()
        c.border = std_border()

        c = ws.cell(row=r, column=3, value=hname)
        c.font = font(9); c.fill = fill; c.alignment = align(h="left", indent=1)
        c.border = std_border()
        ws.row_dimensions[r].height = 16

    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = "2F5496"


# ==================================================
#  ガントチャートシート
# ==================================================
def create_gantt_sheet(wb, tasks):
    ws = wb.create_sheet("ガントチャート")

    # 日付範囲
    all_starts = [t["start"] for t in tasks]
    all_ends   = [t["end"]   for t in tasks]
    min_date   = min(all_starts).replace(day=1)
    max_date   = max(all_ends) + timedelta(days=5)

    dates = []
    d = min_date
    while d <= max_date:
        dates.append(d)
        d += timedelta(days=1)

    # 列レイアウト
    #  A: No.  B: グループ  C: ステータス  D: タスク名  E: 担当者  F: 進捗バー  G~: 日付
    INFO_COLS = 6
    DATE_COL_START = INFO_COLS + 1
    TITLE_ROW  = 1
    MONTH_ROW  = 2
    DAY_ROW    = 3
    DOW_ROW    = 4
    DATA_ROW   = 5

    col_widths = {"A": 4.5, "B": 8, "C": 8, "D": 22, "E": 7, "F": 7}
    for letter, w in col_widths.items():
        ws.column_dimensions[letter].width = w

    # === タイトル行 ===
    last_col = get_column_letter(DATE_COL_START + len(dates) - 1)
    ws.merge_cells(f"A{TITLE_ROW}:{last_col}{TITLE_ROW}")
    c = ws["A1"]
    c.value = PROJECT_TITLE
    c.font = font(16, bold=True, color=HEADER_FG)
    c.fill = make_fill(HEADER_BG)
    c.alignment = align()
    ws.row_dimensions[TITLE_ROW].height = 40

    # === 左側ヘッダー (rows 2-4) ===
    left_headers = ["No.", "グループ", "ステータス", "タスク名", "担当者", "進捗"]
    for ci, lbl in enumerate(left_headers, start=1):
        ws.merge_cells(f"{get_column_letter(ci)}{MONTH_ROW}:{get_column_letter(ci)}{DOW_ROW}")
        c = ws.cell(row=MONTH_ROW, column=ci, value=lbl)
        c.font = font(8, bold=True, color=HEADER_FG)
        c.fill = make_fill(HEADER_BG)
        c.alignment = align()
        c.border = std_border()
        for r in [DAY_ROW, DOW_ROW]:
            ws.cell(row=r, column=ci).fill = make_fill(HEADER_BG)
            ws.cell(row=r, column=ci).border = std_border()

    ws.row_dimensions[MONTH_ROW].height = 20
    ws.row_dimensions[DAY_ROW].height = 16
    ws.row_dimensions[DOW_ROW].height = 14

    # === 月ヘッダー ===
    month_groups = []
    cur_m = cur_y = cur_sc = None
    for i, d in enumerate(dates):
        col = DATE_COL_START + i
        if d.month != cur_m:
            if cur_m is not None:
                month_groups.append((cur_m, cur_y, cur_sc, col - 1))
            cur_m, cur_y, cur_sc = d.month, d.year, col
    if cur_m is not None:
        month_groups.append((cur_m, cur_y, cur_sc, DATE_COL_START + len(dates) - 1))

    MONTH_JP = ["1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月"]
    for (month, year, sc, ec) in month_groups:
        if sc != ec:
            ws.merge_cells(f"{get_column_letter(sc)}{MONTH_ROW}:{get_column_letter(ec)}{MONTH_ROW}")
        c = ws.cell(row=MONTH_ROW, column=sc, value=f"{year}年 {MONTH_JP[month-1]}")
        c.font = font(9, bold=True, color=HEADER_FG)
        c.fill = make_fill(MONTH_BG)
        c.alignment = align()
        c.border = std_border()

    # === 日付ヘッダー + 曜日 ===
    WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]
    for i, d in enumerate(dates):
        col = DATE_COL_START + i
        cl = get_column_letter(col)
        ws.column_dimensions[cl].width = 3.0

        is_we = d.weekday() >= 5
        is_hol = is_holiday(d)
        is_month_first = d.day == 1

        if is_hol:
            day_bg = HOLIDAY_BG
            day_fg = "FFC00000"
            dow_bg = HOLIDAY_BG
            dow_fg = "FFC00000"
        elif is_we:
            day_bg = WEEKEND_BG
            day_fg = "FF595959"
            dow_bg = WEEKEND_BG
            dow_fg = "FF999999"
        else:
            day_bg = "FFD9E1F2"
            day_fg = HEADER_BG
            dow_bg = "FFE8EDF5"
            dow_fg = "FF595959"

        # 日付番号
        c = ws.cell(row=DAY_ROW, column=col, value=d.day)
        c.font = font(7, bold=True, color=day_fg)
        c.fill = make_fill(day_bg)
        c.alignment = align()
        left_style = "medium" if is_month_first else "thin"
        left_color = MONTH_LINE if is_month_first else GRIDLINE
        c.border = border(left=left_style, left_c=left_color)

        # 曜日
        wd_str = WEEKDAY_JP[d.weekday()]
        c = ws.cell(row=DOW_ROW, column=col, value=wd_str)
        c.font = font(6, color=dow_fg)
        c.fill = make_fill(dow_bg)
        c.alignment = align()
        c.border = border(left=left_style, left_c=left_color)

    # === データ行（タスクごと）===
    for task_idx, task in enumerate(tasks):
        row = DATA_ROW + task_idx
        color = TASK_COLORS[task_idx % len(TASK_COLORS)]
        bg = ROW_EVEN if task_idx % 2 == 0 else ROW_ODD
        is_milestone = task["start"] == task["end"]
        progress = task.get("progress", 0) / 100.0

        ws.row_dimensions[row].height = 24

        # No.
        c = ws.cell(row=row, column=1, value=task_idx + 1)
        c.font = font(8, color="FF595959")
        c.fill = make_fill(bg)
        c.alignment = align()
        c.border = std_border()

        # グループ
        group = task.get("group", "")
        gc = GROUP_COLORS.get(group)
        if gc:
            c = ws.cell(row=row, column=2, value=group)
            c.font = font(7, bold=True, color=gc["fg"])
            c.fill = make_fill(gc["bg"])
        else:
            c = ws.cell(row=row, column=2, value=group)
            c.font = font(7, color="FF595959")
            c.fill = make_fill(bg)
        c.alignment = align()
        c.border = std_border()

        # ステータス
        st = task.get("status", "未着手")
        sc_dict = STATUS_COLORS.get(st, STATUS_COLORS["未着手"])
        c = ws.cell(row=row, column=3, value=st)
        c.font = font(7, bold=True, color=sc_dict["fg"])
        c.fill = make_fill(sc_dict["bg"])
        c.alignment = align()
        c.border = std_border()

        # タスク名
        c = ws.cell(row=row, column=4, value=task["name"])
        c.font = font(9, bold=True, color="FF1F3864")
        c.fill = make_fill(bg)
        c.alignment = align(h="left", indent=1)
        c.border = std_border()

        # 担当者
        c = ws.cell(row=row, column=5, value=task.get("assignee", ""))
        c.font = font(8)
        c.fill = make_fill(bg)
        c.alignment = align()
        c.border = std_border()

        # 進捗バー
        c = ws.cell(row=row, column=6, value=f"{int(progress*100)}%")
        if progress >= 1.0:
            c.font = font(8, bold=True, color="FF375623")
            c.fill = make_fill("FFE2EFDA")
        elif progress > 0:
            c.font = font(8, bold=True, color="FF1F4E79")
            c.fill = make_fill("FFDCE6F1")
        else:
            c.font = font(8, color="FF999999")
            c.fill = make_fill(bg)
        c.alignment = align()
        c.border = std_border()

        # === 日付セルの塗りつぶし ===
        task_start_idx = None
        task_end_idx = None
        for i, d in enumerate(dates):
            if d == task["start"]:
                task_start_idx = i
            if d == task["end"]:
                task_end_idx = i

        for i, d in enumerate(dates):
            col = DATE_COL_START + i
            is_we = d.weekday() >= 5
            is_hol = is_holiday(d)
            is_month_first = d.day == 1

            in_range = task["start"] <= d <= task["end"]

            c = ws.cell(row=row, column=col, value="")

            if is_milestone and d == task["start"]:
                c.value = "\u25c6"
                c.font = font(10, bold=True, color=MILESTONE_COLOR)
                c.fill = make_fill(bg)
                c.alignment = align()
            elif in_range and not is_milestone:
                if task_start_idx is not None and task_end_idx is not None:
                    total_cells = task_end_idx - task_start_idx + 1
                    progress_cells = math.floor(progress * total_cells)
                    cell_offset = i - task_start_idx
                    if cell_offset < progress_cells:
                        c.fill = make_fill(color["bar"])
                    else:
                        c.fill = make_fill(color["light"])
                else:
                    c.fill = make_fill(color["bar"])
            else:
                if is_hol:
                    c.fill = make_fill(HOLIDAY_BG)
                elif is_we:
                    c.fill = make_fill(WEEKEND_BG)
                else:
                    c.fill = make_fill(bg)

            # ボーダー
            left_style = "medium" if is_month_first else None
            left_color = MONTH_LINE if is_month_first else GRIDLINE

            if in_range and not is_milestone:
                top_s = Side(style="thin", color=color["light"])
                bot_s = Side(style="thin", color=color["light"])
                l_s = Side(style="thin", color="FFFFFFFF") if d == task["start"] else (
                    Side(style="medium", color=MONTH_LINE) if is_month_first else Side(style=None))
                r_s = Side(style="thin", color="FFFFFFFF") if d == task["end"] else Side(style=None)
                c.border = Border(left=l_s, right=r_s, top=top_s, bottom=bot_s)
            else:
                c.border = border(
                    left=left_style if left_style else "thin",
                    left_c=left_color,
                )

    # === 今日の日付ライン ===
    today = date.today()
    if min_date <= today <= max_date:
        today_idx = (today - min_date).days
        today_col = DATE_COL_START + today_idx

        for r in [DAY_ROW, DOW_ROW]:
            c = ws.cell(row=r, column=today_col)
            c.fill = make_fill(TODAY_RED)
            c.font = font(c.font.size or 7, bold=True, color="FFFFFFFF")

        for task_idx in range(len(tasks)):
            row = DATA_ROW + task_idx
            c = ws.cell(row=row, column=today_col)
            c.border = Border(
                left=Side(style="medium", color=TODAY_RED),
                right=Side(style="medium", color=TODAY_RED),
                top=Side(style="thin", color=GRIDLINE),
                bottom=Side(style="thin", color=GRIDLINE),
            )

    # === 凡例 ===
    lr = DATA_ROW + len(tasks) + 1
    ws.row_dimensions[lr].height = 6

    lr += 1
    ws.cell(row=lr, column=1, value="凡例").font = font(9, bold=True, color=HEADER_BG[2:])
    ws.row_dimensions[lr].height = 18

    legend_items = [
        ("濃色", TASK_COLORS[0]["bar"], "進捗済み期間"),
        ("淡色", TASK_COLORS[0]["light"], "未進捗期間"),
        ("  ",  WEEKEND_BG, "土曜・日曜"),
        ("  ",  HOLIDAY_BG, "祝日"),
        ("\u25c6",  None, "マイルストーン"),
        ("|",   None, "本日ライン（赤）"),
    ]

    lr += 1
    for idx, (sym, color_hex, desc) in enumerate(legend_items):
        r = lr + idx
        if color_hex:
            c = ws.cell(row=r, column=1, value="  ")
            c.fill = make_fill(color_hex)
            c.border = std_border()
        else:
            c = ws.cell(row=r, column=1, value=sym)
            c.font = font(9, bold=True, color="FFFF4444" if sym in ("\u25c6", "|") else "FF333333")
        ws.merge_cells(f"B{r}:C{r}")
        c = ws.cell(row=r, column=2, value=desc)
        c.font = font(8, color="FF595959")
        ws.row_dimensions[r].height = 15

    # === シート設定 ===
    ws.freeze_panes = f"{get_column_letter(DATE_COL_START)}{DATA_ROW}"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "4472C4"

    # 印刷設定
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{MONTH_ROW}:{DOW_ROW}"
    ws.print_title_cols = f"A:{get_column_letter(INFO_COLS)}"
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4


# ==================================================
#  使い方シート
# ==================================================
def create_help_sheet(wb):
    ws = wb.create_sheet("使い方")

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "使い方ガイド"
    c.font = font(16, bold=True, color=HEADER_FG)
    c.fill = make_fill(HEADER_BG)
    c.alignment = align()
    ws.row_dimensions[1].height = 36
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 60

    content = [
        ("■ ガントチャートの更新方法", [
            "1. 「入力」シートのタスクデータを確認・変更します",
            "2. 「ガントチャート生成.py」の TASKS リストを同様に編集します",
            "3. ターミナルで以下を実行します:",
            "     pip install openpyxl",
            "     python ガントチャート生成.py",
            "4. 同じフォルダに新しい「ガントチャート.xlsx」が生成されます",
        ]),
        ("■ CLIオプション", [
            "  python ガントチャート生成.py -o 出力ファイル.xlsx",
            "  python ガントチャート生成.py --import tasks.json",
            "  python ガントチャート生成.py --import tasks.csv",
            "  python ガントチャート生成.py --export tasks.json",
        ]),
        ("■ タスクデータの書き方", [
            '  {"id": 1, "name": "タスク名",',
            '   "start": date(2026, 4, 1),  # 年, 月, 日',
            '   "end": date(2026, 4, 7),',
            '   "assignee": "担当者名",',
            '   "status": "未着手",  # 完了/進行中/未着手/遅延/中断',
            '   "priority": "高",    # 最高/高/中/低',
            '   "progress": 60,      # 0〜100の整数（%）',
            '   "group": "企画",     # グループ名（任意）',
            '   "depends_on": [1],   # 依存タスクID（任意）',
            '   "note": "備考テキスト"},',
        ]),
        ("■ JSON形式でのタスク定義", [
            '  {"tasks": [',
            '    {"id": 1, "name": "要件定義", "start": "2026-04-01",',
            '     "end": "2026-04-07", "assignee": "田中",',
            '     "status": "完了", "priority": "高", "progress": 100,',
            '     "group": "企画", "note": "備考"}',
            '  ]}',
        ]),
        ("■ マイルストーンの設定方法", [
            "開始日と終了日を同じ日付にするとマイルストーン（\u25c6）として表示されます",
            '  例: {"name": "★ リリース", "start": date(2026, 5, 28), "end": date(2026, 5, 28), ...}',
        ]),
        ("■ 依存関係について", [
            "depends_on に依存先タスクのIDリストを指定できます",
            "入力シートの備考欄に依存関係が自動表示されます",
            '  例: "depends_on": [1, 2]  → タスク#1と#2に依存',
        ]),
        ("■ 色の説明", [
            "・ 濃い色のバー → 進捗済み部分",
            "・ 薄い色のバー → 残作業部分",
            "・ 薄いグレー → 土曜・日曜",
            "・ 薄いピンク → 祝日",
            "・ 赤い縦線 → 本日の日付",
        ]),
        ("■ 祝日について", [
            "スクリプト内の HOLIDAYS 辞書に2025〜2027年の祝日が定義されています",
            "年度が変わる場合はこの辞書を更新してください",
            "稼働日数は土日祝を除いた営業日ベースで自動計算されます",
        ]),
        ("■ 印刷設定", [
            "ガントチャートシートはA3横向き・縮小印刷に設定済みです",
            "タスク名と日付ヘッダーは印刷時にも各ページに表示されます",
        ]),
    ]

    row = 3
    for section_title, lines in content:
        c = ws.cell(row=row, column=2, value=section_title)
        c.font = font(11, bold=True, color=HEADER_BG[2:])
        ws.row_dimensions[row].height = 24
        row += 1
        for line in lines:
            c = ws.cell(row=row, column=2, value=line)
            c.font = font(9, color="FF333333")
            c.alignment = align(h="left", indent=1, wrap=True)
            ws.row_dimensions[row].height = 17
            row += 1
        row += 1

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "70AD47"


# ==================================================
#  CLI引数パース
# ==================================================
def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description="ガントチャート生成スクリプト（JTC業務品質版）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "-o", "--output",
        default=OUTPUT_FILE,
        help=f"出力ファイル名 (デフォルト: {OUTPUT_FILE})",
    )
    parser.add_argument(
        "--import", dest="import_file",
        help="タスクデータのインポート元（.json または .csv）",
    )
    parser.add_argument(
        "--export", dest="export_file",
        help="現在のタスクデータをJSON形式でエクスポート",
    )
    parser.add_argument(
        "--validate-only", action="store_true",
        help="タスクデータのバリデーションのみ実行",
    )
    return parser.parse_args(argv)


# ==================================================
#  メイン
# ==================================================
def generate_gantt(tasks, output_file):
    """ガントチャートExcelを生成して保存する。"""
    validate_tasks(tasks)
    wb = Workbook()
    create_input_sheet(wb, tasks)
    create_gantt_sheet(wb, tasks)
    create_help_sheet(wb)
    wb.save(output_file)
    return wb


def main(argv=None):
    args = parse_args(argv)

    tasks = TASKS

    # インポート
    if args.import_file:
        ext = os.path.splitext(args.import_file)[1].lower()
        if ext == ".json":
            tasks = import_tasks_from_json(args.import_file)
        elif ext == ".csv":
            tasks = import_tasks_from_csv(args.import_file)
        else:
            print(f"エラー: 未対応のファイル形式です: {ext}")
            sys.exit(1)
        print(f"インポート完了: {args.import_file} ({len(tasks)}件)")

    # エクスポート
    if args.export_file:
        export_tasks_to_json(tasks, args.export_file)
        print(f"エクスポート完了: {args.export_file}")
        if not args.validate_only and not args.import_file:
            pass  # 続行してExcelも生成

    # バリデーション
    try:
        validate_tasks(tasks)
        if args.validate_only:
            print("バリデーション成功: すべてのタスクデータが正常です")
            return
    except TaskValidationError as e:
        print(f"バリデーションエラー:\n{e}", file=sys.stderr)
        sys.exit(1)

    # 生成
    generate_gantt(tasks, args.output)
    print(f"ガントチャートを生成しました: {args.output}")
    print(f"   タスク数: {len(tasks)}")

    # グループ集計
    groups = []
    seen = set()
    for t in tasks:
        g = t.get("group", "")
        if g and g not in seen:
            groups.append(g)
            seen.add(g)
    if groups:
        print(f"   グループ: {', '.join(groups)}")

    holiday_years = set()
    for t in tasks:
        holiday_years.add(t["start"].year)
        holiday_years.add(t["end"].year)
    relevant = sum(1 for h in HOLIDAYS if h.year in holiday_years)
    print(f"   祝日数:   {relevant}")
    print(f"   シート:   入力 / ガントチャート / 使い方")


if __name__ == "__main__":
    main()
