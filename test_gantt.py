#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ガントチャート生成スクリプトのテストスイート
"""

import csv
import json
import os
import tempfile

import pytest

# テスト対象のインポート（ファイル名が日本語のためimportlibを使用）
import importlib.util
SCRIPT_PATH = os.path.join(os.path.dirname(__file__), "ガントチャート生成.py")
spec = importlib.util.spec_from_file_location("gantt", SCRIPT_PATH)
gantt = importlib.util.module_from_spec(spec)
spec.loader.exec_module(gantt)

from datetime import date, timedelta
from openpyxl import load_workbook


# ==================================================
# ユーティリティ関数のテスト
# ==================================================
class TestHolidays:
    def test_known_holiday(self):
        assert gantt.is_holiday(date(2026, 1, 1)) is True  # 元日
        assert gantt.is_holiday(date(2026, 5, 5)) is True  # こどもの日

    def test_non_holiday(self):
        assert gantt.is_holiday(date(2026, 4, 1)) is False  # 普通の平日

    def test_2025_holidays_exist(self):
        assert gantt.is_holiday(date(2025, 1, 1)) is True
        assert gantt.is_holiday(date(2025, 2, 24)) is True  # 振替休日

    def test_2027_holidays_exist(self):
        assert gantt.is_holiday(date(2027, 1, 1)) is True
        assert gantt.is_holiday(date(2027, 3, 22)) is True  # 振替休日

    def test_holidays_2026_backward_compat(self):
        """HOLIDAYS_2026エイリアスが正しく2026年のみ含むことを確認"""
        for d in gantt.HOLIDAYS_2026:
            assert d.year == 2026


class TestNonWorking:
    def test_saturday(self):
        assert gantt.is_non_working(date(2026, 4, 4)) is True  # 土曜

    def test_sunday(self):
        assert gantt.is_non_working(date(2026, 4, 5)) is True  # 日曜

    def test_weekday(self):
        assert gantt.is_non_working(date(2026, 4, 1)) is False  # 水曜

    def test_holiday_weekday(self):
        assert gantt.is_non_working(date(2026, 4, 29)) is True  # 昭和の日（水曜）


class TestCountWorkingDays:
    def test_single_day_weekday(self):
        assert gantt.count_working_days(date(2026, 4, 1), date(2026, 4, 1)) == 1

    def test_single_day_weekend(self):
        assert gantt.count_working_days(date(2026, 4, 4), date(2026, 4, 4)) == 0

    def test_full_week(self):
        # 月〜金 = 5日間
        assert gantt.count_working_days(date(2026, 4, 6), date(2026, 4, 10)) == 5

    def test_week_with_weekend(self):
        # 月〜日 = 5稼働日
        assert gantt.count_working_days(date(2026, 4, 6), date(2026, 4, 12)) == 5

    def test_golden_week(self):
        # 5/3-5/6は祝日（憲法記念日、みどりの日、こどもの日、振替休日）
        # 5/1(金)=稼働、5/2(土)=休、5/3(日祝)=休、5/4(月祝)=休、5/5(火祝)=休、5/6(水振替)=休、5/7(木)=稼働
        result = gantt.count_working_days(date(2026, 5, 1), date(2026, 5, 7))
        assert result == 2  # 5/1と5/7のみ


# ==================================================
# バリデーションのテスト
# ==================================================
class TestValidation:
    def _make_task(self, **overrides):
        base = {
            "id": 1,
            "name": "テストタスク",
            "start": date(2026, 4, 1),
            "end": date(2026, 4, 7),
            "assignee": "テスト",
            "status": "未着手",
            "priority": "中",
            "progress": 0,
            "note": "",
        }
        base.update(overrides)
        return base

    def test_valid_task(self):
        gantt.validate_task(self._make_task(), 0)

    def test_missing_name(self):
        with pytest.raises(gantt.TaskValidationError, match="name"):
            gantt.validate_task(self._make_task(name=""), 0)

    def test_missing_start(self):
        task = self._make_task()
        del task["start"]
        with pytest.raises(gantt.TaskValidationError, match="start"):
            gantt.validate_task(task, 0)

    def test_missing_end(self):
        task = self._make_task()
        del task["end"]
        with pytest.raises(gantt.TaskValidationError, match="end"):
            gantt.validate_task(task, 0)

    def test_start_after_end(self):
        with pytest.raises(gantt.TaskValidationError, match="開始日.*終了日"):
            gantt.validate_task(
                self._make_task(start=date(2026, 4, 10), end=date(2026, 4, 1)), 0
            )

    def test_invalid_status(self):
        with pytest.raises(gantt.TaskValidationError, match="ステータス"):
            gantt.validate_task(self._make_task(status="不明"), 0)

    def test_invalid_priority(self):
        with pytest.raises(gantt.TaskValidationError, match="優先度"):
            gantt.validate_task(self._make_task(priority="超高"), 0)

    def test_progress_negative(self):
        with pytest.raises(gantt.TaskValidationError, match="progress"):
            gantt.validate_task(self._make_task(progress=-10), 0)

    def test_progress_over_100(self):
        with pytest.raises(gantt.TaskValidationError, match="progress"):
            gantt.validate_task(self._make_task(progress=150), 0)

    def test_valid_statuses(self):
        for status in gantt.VALID_STATUSES:
            gantt.validate_task(self._make_task(status=status), 0)

    def test_valid_priorities(self):
        for priority in gantt.VALID_PRIORITIES:
            gantt.validate_task(self._make_task(priority=priority), 0)

    def test_milestone_task(self):
        """マイルストーン（開始=終了）は正常"""
        gantt.validate_task(
            self._make_task(start=date(2026, 5, 28), end=date(2026, 5, 28)), 0
        )


class TestValidateTasks:
    def _make_tasks(self):
        return [
            {"id": 1, "name": "タスクA", "start": date(2026, 4, 1), "end": date(2026, 4, 7),
             "status": "未着手", "priority": "中", "progress": 0},
            {"id": 2, "name": "タスクB", "start": date(2026, 4, 8), "end": date(2026, 4, 14),
             "status": "未着手", "priority": "中", "progress": 0, "depends_on": [1]},
        ]

    def test_valid_tasks(self):
        gantt.validate_tasks(self._make_tasks())

    def test_duplicate_id(self):
        tasks = self._make_tasks()
        tasks[1]["id"] = 1  # 重複ID
        with pytest.raises(gantt.TaskValidationError, match="重複"):
            gantt.validate_tasks(tasks)

    def test_missing_dependency(self):
        tasks = self._make_tasks()
        tasks[1]["depends_on"] = [99]  # 存在しないID
        with pytest.raises(gantt.TaskValidationError, match="依存先ID 99"):
            gantt.validate_tasks(tasks)


# ==================================================
# JSON/CSVインポート・エクスポートのテスト
# ==================================================
class TestJsonImportExport:
    def test_roundtrip(self, tmp_path):
        """エクスポートしたJSONをインポートして同一内容になることを確認"""
        original_tasks = [
            {"id": 1, "name": "テスト", "start": date(2026, 4, 1), "end": date(2026, 4, 7),
             "assignee": "太郎", "status": "進行中", "priority": "高", "progress": 50,
             "note": "メモ", "group": "企画", "depends_on": []},
        ]
        json_path = str(tmp_path / "tasks.json")
        gantt.export_tasks_to_json(original_tasks, json_path)

        imported = gantt.import_tasks_from_json(json_path)
        assert len(imported) == 1
        assert imported[0]["name"] == "テスト"
        assert imported[0]["start"] == date(2026, 4, 1)
        assert imported[0]["end"] == date(2026, 4, 7)
        assert imported[0]["progress"] == 50

    def test_import_with_depends_on_string(self, tmp_path):
        """depends_onが文字列の場合でもパースできる"""
        data = {"tasks": [
            {"id": 1, "name": "A", "start": "2026-04-01", "end": "2026-04-07",
             "progress": 0},
            {"id": 2, "name": "B", "start": "2026-04-08", "end": "2026-04-14",
             "progress": 0, "depends_on": "1"},
        ]}
        json_path = str(tmp_path / "tasks.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f)

        imported = gantt.import_tasks_from_json(json_path)
        assert imported[1]["depends_on"] == [1]

    def test_import_flat_array(self, tmp_path):
        """トップレベルが配列のJSONも対応"""
        data = [
            {"id": 1, "name": "A", "start": "2026-04-01", "end": "2026-04-07", "progress": 0},
        ]
        json_path = str(tmp_path / "tasks.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f)

        imported = gantt.import_tasks_from_json(json_path)
        assert len(imported) == 1


class TestCsvImport:
    def test_basic_import(self, tmp_path):
        csv_path = str(tmp_path / "tasks.csv")
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=[
                "id", "name", "start", "end", "assignee", "status", "priority",
                "progress", "group", "depends_on", "note"
            ])
            writer.writeheader()
            writer.writerow({
                "id": "1", "name": "要件定義", "start": "2026-04-01", "end": "2026-04-07",
                "assignee": "田中", "status": "完了", "priority": "高",
                "progress": "100", "group": "企画", "depends_on": "", "note": "完了済み"
            })
            writer.writerow({
                "id": "2", "name": "基本設計", "start": "2026-04-08", "end": "2026-04-14",
                "assignee": "鈴木", "status": "進行中", "priority": "高",
                "progress": "60", "group": "企画", "depends_on": "1", "note": ""
            })

        tasks = gantt.import_tasks_from_csv(csv_path)
        assert len(tasks) == 2
        assert tasks[0]["name"] == "要件定義"
        assert tasks[0]["start"] == date(2026, 4, 1)
        assert tasks[0]["progress"] == 100
        assert tasks[1]["depends_on"] == [1]


# ==================================================
# Excel生成のテスト
# ==================================================
class TestExcelGeneration:
    @pytest.fixture
    def sample_tasks(self):
        return [
            {"id": 1, "name": "タスクA", "start": date(2026, 4, 1), "end": date(2026, 4, 7),
             "assignee": "田中", "status": "完了", "priority": "高", "progress": 100,
             "group": "企画", "note": "テスト"},
            {"id": 2, "name": "タスクB", "start": date(2026, 4, 8), "end": date(2026, 4, 14),
             "assignee": "鈴木", "status": "進行中", "priority": "中", "progress": 50,
             "group": "開発", "depends_on": [1], "note": ""},
            {"id": 3, "name": "★ マイルストーン", "start": date(2026, 4, 15), "end": date(2026, 4, 15),
             "assignee": "全員", "status": "未着手", "priority": "最高", "progress": 0,
             "group": "リリース", "depends_on": [2], "note": "マイルストーン"},
        ]

    def test_generate_creates_file(self, sample_tasks, tmp_path):
        output = str(tmp_path / "test.xlsx")
        gantt.generate_gantt(sample_tasks, output)
        assert os.path.exists(output)

    def test_has_three_sheets(self, sample_tasks, tmp_path):
        output = str(tmp_path / "test.xlsx")
        gantt.generate_gantt(sample_tasks, output)
        wb = load_workbook(output)
        assert wb.sheetnames == ["入力", "ガントチャート", "使い方"]

    def test_input_sheet_title(self, sample_tasks, tmp_path):
        output = str(tmp_path / "test.xlsx")
        gantt.generate_gantt(sample_tasks, output)
        wb = load_workbook(output)
        ws = wb["入力"]
        assert gantt.PROJECT_TITLE in str(ws["A1"].value)

    def test_input_sheet_task_count(self, sample_tasks, tmp_path):
        output = str(tmp_path / "test.xlsx")
        gantt.generate_gantt(sample_tasks, output)
        wb = load_workbook(output)
        ws = wb["入力"]
        # タスクはrow 5から始まるので3タスク = row 5,6,7
        assert ws.cell(row=5, column=5).value == "タスクA"
        assert ws.cell(row=6, column=5).value == "タスクB"
        assert ws.cell(row=7, column=5).value == "★ マイルストーン"

    def test_gantt_sheet_has_dates(self, sample_tasks, tmp_path):
        output = str(tmp_path / "test.xlsx")
        gantt.generate_gantt(sample_tasks, output)
        wb = load_workbook(output)
        ws = wb["ガントチャート"]
        # DATE_COL_START = 7, 最初の日付は4月1日のある月の1日
        # row 3が日付行
        assert ws.cell(row=3, column=7).value == 1  # 月の最初の日

    def test_gantt_milestone_marker(self, sample_tasks, tmp_path):
        output = str(tmp_path / "test.xlsx")
        gantt.generate_gantt(sample_tasks, output)
        wb = load_workbook(output)
        ws = wb["ガントチャート"]
        # マイルストーン（タスク3 = row 7）のセルに◆があるか
        found_diamond = False
        for col in range(7, ws.max_column + 1):
            if ws.cell(row=7, column=col).value == "\u25c6":
                found_diamond = True
                break
        assert found_diamond, "マイルストーンマーカー◆が見つかりません"

    def test_group_summary_exists(self, sample_tasks, tmp_path):
        output = str(tmp_path / "test.xlsx")
        gantt.generate_gantt(sample_tasks, output)
        wb = load_workbook(output)
        ws = wb["入力"]
        # グループ別サマリーがあるかチェック
        found_summary = False
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val and "グループ別サマリー" in str(val):
                found_summary = True
                break
        assert found_summary, "グループ別サマリーが見つかりません"

    def test_dependency_note(self, sample_tasks, tmp_path):
        """依存関係が備考欄に表示されることを確認"""
        output = str(tmp_path / "test.xlsx")
        gantt.generate_gantt(sample_tasks, output)
        wb = load_workbook(output)
        ws = wb["入力"]
        # タスクB (row 6) の備考欄 (column 12) に依存情報がある
        note = str(ws.cell(row=6, column=12).value)
        assert "依存" in note and "#1" in note

    def test_help_sheet_content(self, sample_tasks, tmp_path):
        output = str(tmp_path / "test.xlsx")
        gantt.generate_gantt(sample_tasks, output)
        wb = load_workbook(output)
        ws = wb["使い方"]
        assert "使い方ガイド" in str(ws["A1"].value)


class TestExcelWithDifferentStatuses:
    """全ステータスの色分けが正しく生成されることを確認"""

    @pytest.fixture
    def all_status_tasks(self):
        tasks = []
        for i, status in enumerate(gantt.VALID_STATUSES):
            tasks.append({
                "id": i + 1,
                "name": f"タスク_{status}",
                "start": date(2026, 4, 1 + i),
                "end": date(2026, 4, 5 + i),
                "assignee": "テスト",
                "status": status,
                "priority": "中",
                "progress": 50 if status == "進行中" else (100 if status == "完了" else 0),
                "note": "",
            })
        return tasks

    def test_all_statuses_generate(self, all_status_tasks, tmp_path):
        output = str(tmp_path / "test.xlsx")
        gantt.generate_gantt(all_status_tasks, output)
        assert os.path.exists(output)


# ==================================================
# CLI引数のテスト
# ==================================================
class TestParseArgs:
    def test_defaults(self):
        args = gantt.parse_args([])
        assert args.output == gantt.OUTPUT_FILE
        assert args.import_file is None
        assert args.export_file is None
        assert args.validate_only is False

    def test_output(self):
        args = gantt.parse_args(["-o", "custom.xlsx"])
        assert args.output == "custom.xlsx"

    def test_import(self):
        args = gantt.parse_args(["--import", "tasks.json"])
        assert args.import_file == "tasks.json"

    def test_export(self):
        args = gantt.parse_args(["--export", "out.json"])
        assert args.export_file == "out.json"

    def test_validate_only(self):
        args = gantt.parse_args(["--validate-only"])
        assert args.validate_only is True


# ==================================================
# 統合テスト
# ==================================================
class TestIntegration:
    def test_main_with_default_tasks(self, tmp_path):
        """デフォルトタスクでのメイン実行"""
        output = str(tmp_path / "output.xlsx")
        gantt.main(["-o", output])
        assert os.path.exists(output)
        wb = load_workbook(output)
        assert len(wb.sheetnames) == 3

    def test_main_with_json_import(self, tmp_path):
        """JSONインポートからの生成"""
        json_path = os.path.join(os.path.dirname(__file__), "sample_tasks.json")
        output = str(tmp_path / "output.xlsx")
        gantt.main(["--import", json_path, "-o", output])
        assert os.path.exists(output)

    def test_main_validate_only(self, capsys):
        """バリデーションのみモード"""
        gantt.main(["--validate-only"])
        captured = capsys.readouterr()
        assert "バリデーション成功" in captured.out

    def test_main_export(self, tmp_path):
        """エクスポートモード"""
        export_path = str(tmp_path / "export.json")
        output = str(tmp_path / "output.xlsx")
        gantt.main(["--export", export_path, "-o", output])
        assert os.path.exists(export_path)
        with open(export_path, encoding="utf-8") as f:
            data = json.load(f)
        assert "tasks" in data
        assert len(data["tasks"]) == len(gantt.TASKS)

    def test_full_pipeline_json(self, tmp_path):
        """JSON経由のフルパイプライン: エクスポート → インポート → 生成"""
        export_path = str(tmp_path / "exported.json")
        gantt.export_tasks_to_json(gantt.TASKS, export_path)

        imported = gantt.import_tasks_from_json(export_path)
        assert len(imported) == len(gantt.TASKS)

        output = str(tmp_path / "regenerated.xlsx")
        gantt.generate_gantt(imported, output)
        assert os.path.exists(output)

        wb = load_workbook(output)
        assert wb.sheetnames == ["入力", "ガントチャート", "使い方"]


# ==================================================
# エッジケースのテスト
# ==================================================
class TestEdgeCases:
    def test_single_task(self, tmp_path):
        """タスク1件での生成"""
        tasks = [
            {"id": 1, "name": "単独タスク", "start": date(2026, 4, 1), "end": date(2026, 4, 3),
             "assignee": "太郎", "status": "未着手", "priority": "中", "progress": 0, "note": ""},
        ]
        output = str(tmp_path / "test.xlsx")
        gantt.generate_gantt(tasks, output)
        assert os.path.exists(output)

    def test_many_tasks(self, tmp_path):
        """大量タスク（20件）での生成"""
        tasks = []
        for i in range(20):
            tasks.append({
                "id": i + 1,
                "name": f"タスク{i+1}",
                "start": date(2026, 4, 1) + timedelta(days=i * 3),
                "end": date(2026, 4, 1) + timedelta(days=i * 3 + 5),
                "assignee": "担当者",
                "status": "未着手",
                "priority": "中",
                "progress": 0,
                "note": "",
            })
        output = str(tmp_path / "test.xlsx")
        gantt.generate_gantt(tasks, output)
        assert os.path.exists(output)

    def test_task_spanning_months(self, tmp_path):
        """月をまたぐタスクの生成"""
        tasks = [
            {"id": 1, "name": "月跨ぎタスク", "start": date(2026, 4, 25), "end": date(2026, 5, 10),
             "assignee": "太郎", "status": "進行中", "priority": "高", "progress": 40, "note": ""},
        ]
        output = str(tmp_path / "test.xlsx")
        gantt.generate_gantt(tasks, output)
        assert os.path.exists(output)

    def test_100_percent_progress(self, tmp_path):
        """進捗100%のタスク"""
        tasks = [
            {"id": 1, "name": "完了タスク", "start": date(2026, 4, 1), "end": date(2026, 4, 7),
             "assignee": "太郎", "status": "完了", "priority": "中", "progress": 100, "note": ""},
        ]
        output = str(tmp_path / "test.xlsx")
        gantt.generate_gantt(tasks, output)
        assert os.path.exists(output)

    def test_task_without_optional_fields(self, tmp_path):
        """オプションフィールドなしのタスク"""
        tasks = [
            {"name": "最小タスク", "start": date(2026, 4, 1), "end": date(2026, 4, 3)},
        ]
        output = str(tmp_path / "test.xlsx")
        gantt.generate_gantt(tasks, output)
        assert os.path.exists(output)
